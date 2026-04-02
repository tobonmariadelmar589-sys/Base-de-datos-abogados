import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from tkcalendar import DateEntry
import sqlite3, os, re, io

DB_PATH = os.path.join(os.path.dirname(__file__), "lawfirm.db")

def get_conn():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn

def init_db():
    with get_conn() as conn:
        conn.executescript("""
        CREATE TABLE IF NOT EXISTS clientes (
            id_cliente INTEGER PRIMARY KEY AUTOINCREMENT,
            codigo TEXT NOT NULL UNIQUE, tipo TEXT, nombre TEXT NOT NULL,
            documento TEXT NOT NULL UNIQUE, direccion TEXT, telefono TEXT,
            correo TEXT, fecha_contacto TEXT, referido_por TEXT,
            clasificacion TEXT, foto BLOB, fecha_registro TEXT DEFAULT (date('now'))
        );
        CREATE TABLE IF NOT EXISTS abogados (
            id_abogado INTEGER PRIMARY KEY AUTOINCREMENT,
            num_colegiatura TEXT NOT NULL UNIQUE, nombres TEXT NOT NULL,
            apellidos TEXT NOT NULL, especialidad TEXT, anios_exp INTEGER,
            formacion TEXT, idiomas TEXT, tarifa_hora REAL,
            disponibilidad TEXT DEFAULT 'Disponible', foto BLOB,
            fecha_registro TEXT DEFAULT (date('now'))
        );
        CREATE TABLE IF NOT EXISTS casos (
            id_caso INTEGER PRIMARY KEY AUTOINCREMENT,
            numero_caso TEXT NOT NULL UNIQUE, titulo TEXT NOT NULL,
            tipo_caso TEXT, rama_derecho TEXT, fecha_apertura TEXT,
            id_cliente INTEGER REFERENCES clientes(id_cliente),
            contraparte TEXT, juzgado TEXT, num_expediente TEXT,
            id_abogado_principal INTEGER REFERENCES abogados(id_abogado),
            estado TEXT DEFAULT 'Abierto', fecha_conclusion TEXT,
            fecha_registro TEXT DEFAULT (date('now'))
        );
        CREATE TABLE IF NOT EXISTS audiencias (
            id_audiencia INTEGER PRIMARY KEY AUTOINCREMENT,
            codigo TEXT NOT NULL UNIQUE, tipo TEXT,
            id_caso INTEGER REFERENCES casos(id_caso),
            fecha_hora TEXT, duracion_estimada TEXT, lugar TEXT,
            participantes_int TEXT, participantes_ext TEXT, proposito TEXT,
            resultado_esperado TEXT, resultado_real TEXT,
            fecha_registro TEXT DEFAULT (date('now'))
        );
        CREATE VIEW IF NOT EXISTS v_casos AS
            SELECT c.id_caso, c.numero_caso, c.titulo, c.tipo_caso,
                   c.rama_derecho, c.fecha_apertura,
                   cl.nombre AS cliente, c.contraparte,
                   a.nombres||' '||a.apellidos AS abogado_principal,
                   c.estado, c.fecha_conclusion
            FROM casos c
            LEFT JOIN clientes cl ON c.id_cliente = cl.id_cliente
            LEFT JOIN abogados a  ON c.id_abogado_principal = a.id_abogado;
        CREATE VIEW IF NOT EXISTS v_audiencias AS
            SELECT au.id_audiencia, au.codigo, au.tipo,
                   ca.numero_caso, ca.titulo AS caso_titulo,
                   au.fecha_hora, au.duracion_estimada, au.lugar,
                   au.resultado_esperado, au.resultado_real
            FROM audiencias au
            LEFT JOIN casos ca ON au.id_caso = ca.id_caso;
        INSERT OR IGNORE INTO clientes(codigo,tipo,nombre,documento,direccion,telefono,correo,fecha_contacto,referido_por,clasificacion)
        VALUES('CLI-001','Empresa','Constructora Andina S.A.S','900123456-7','Calle 45 #12-30','604-3214567','contacto@constructoraandina.com','2025-01-15','Dr. Ramirez','Cliente VIP');
        INSERT OR IGNORE INTO abogados(num_colegiatura,nombres,apellidos,especialidad,anios_exp,formacion,idiomas,tarifa_hora,disponibilidad)
        VALUES('COL-20456','Carlos Andres','Morales Quintero','Mercantil',12,'Magister Derecho Comercial','Espanol, Ingles',250000,'Disponible');
        INSERT OR IGNORE INTO casos(numero_caso,titulo,tipo_caso,rama_derecho,fecha_apertura,id_cliente,contraparte,juzgado,num_expediente,id_abogado_principal,estado,fecha_conclusion)
        VALUES('CASO-2025-001','Disputa contractual por incumplimiento','Incumplimiento','Civil','2025-02-20',1,'Inmobiliaria del Norte','Juzgado 3 Civil','EXP-0892',1,'En proceso','2025-11-30');
        INSERT OR IGNORE INTO audiencias(codigo,tipo,id_caso,fecha_hora,duracion_estimada,lugar,participantes_int,participantes_ext,proposito,resultado_esperado)
        VALUES('AUD-001','Audiencia Preliminar',1,'2025-04-10 09:00','2 horas','Juzgado 3 Civil Sala 204','Carlos Morales','Rep. Inmobiliaria','Presentacion de pruebas','Admision de pruebas');
        """)

# Stored procedures
def sp_insert_cliente(d):
    with get_conn() as c: c.execute("INSERT INTO clientes(codigo,tipo,nombre,documento,direccion,telefono,correo,fecha_contacto,referido_por,clasificacion,foto) VALUES(:codigo,:tipo,:nombre,:documento,:direccion,:telefono,:correo,:fecha_contacto,:referido_por,:clasificacion,:foto)", d)
def sp_update_cliente(d):
    with get_conn() as c: c.execute("UPDATE clientes SET tipo=:tipo,nombre=:nombre,documento=:documento,direccion=:direccion,telefono=:telefono,correo=:correo,fecha_contacto=:fecha_contacto,referido_por=:referido_por,clasificacion=:clasificacion,foto=:foto WHERE codigo=:codigo", d)
def sp_delete_cliente(codigo):
    with get_conn() as c: c.execute("DELETE FROM clientes WHERE codigo=?", (codigo,))
def sp_get_all_clientes(f=None):
    sql="SELECT * FROM clientes WHERE 1=1"; p=[]
    if f:
        if f.get("tipo"): sql+=" AND tipo=?"; p.append(f["tipo"])
        if f.get("fecha_desde"): sql+=" AND fecha_contacto>=?"; p.append(f["fecha_desde"])
        if f.get("fecha_hasta"): sql+=" AND fecha_contacto<=?"; p.append(f["fecha_hasta"])
        if f.get("clasificacion"): sql+=" AND clasificacion=?"; p.append(f["clasificacion"])
    with get_conn() as c: return c.execute(sql+" ORDER BY nombre", p).fetchall()
def sp_get_cliente(codigo):
    with get_conn() as c: return c.execute("SELECT * FROM clientes WHERE codigo=?", (codigo,)).fetchone()

def sp_insert_abogado(d):
    with get_conn() as c: c.execute("INSERT INTO abogados(num_colegiatura,nombres,apellidos,especialidad,anios_exp,formacion,idiomas,tarifa_hora,disponibilidad,foto) VALUES(:num_colegiatura,:nombres,:apellidos,:especialidad,:anios_exp,:formacion,:idiomas,:tarifa_hora,:disponibilidad,:foto)", d)
def sp_update_abogado(d):
    with get_conn() as c: c.execute("UPDATE abogados SET nombres=:nombres,apellidos=:apellidos,especialidad=:especialidad,anios_exp=:anios_exp,formacion=:formacion,idiomas=:idiomas,tarifa_hora=:tarifa_hora,disponibilidad=:disponibilidad,foto=:foto WHERE num_colegiatura=:num_colegiatura", d)
def sp_delete_abogado(col):
    with get_conn() as c: c.execute("DELETE FROM abogados WHERE num_colegiatura=?", (col,))
def sp_get_all_abogados(f=None):
    sql="SELECT * FROM abogados WHERE 1=1"; p=[]
    if f:
        if f.get("especialidad"): sql+=" AND especialidad=?"; p.append(f["especialidad"])
        if f.get("disponibilidad"): sql+=" AND disponibilidad=?"; p.append(f["disponibilidad"])
    with get_conn() as c: return c.execute(sql+" ORDER BY apellidos", p).fetchall()
def sp_get_abogado(col):
    with get_conn() as c: return c.execute("SELECT * FROM abogados WHERE num_colegiatura=?", (col,)).fetchone()

def sp_insert_caso(d):
    with get_conn() as c: c.execute("INSERT INTO casos(numero_caso,titulo,tipo_caso,rama_derecho,fecha_apertura,id_cliente,contraparte,juzgado,num_expediente,id_abogado_principal,estado,fecha_conclusion) VALUES(:numero_caso,:titulo,:tipo_caso,:rama_derecho,:fecha_apertura,:id_cliente,:contraparte,:juzgado,:num_expediente,:id_abogado_principal,:estado,:fecha_conclusion)", d)
def sp_update_caso(d):
    with get_conn() as c: c.execute("UPDATE casos SET titulo=:titulo,tipo_caso=:tipo_caso,rama_derecho=:rama_derecho,fecha_apertura=:fecha_apertura,id_cliente=:id_cliente,contraparte=:contraparte,juzgado=:juzgado,num_expediente=:num_expediente,id_abogado_principal=:id_abogado_principal,estado=:estado,fecha_conclusion=:fecha_conclusion WHERE numero_caso=:numero_caso", d)
def sp_delete_caso(num):
    with get_conn() as c: c.execute("DELETE FROM casos WHERE numero_caso=?", (num,))
def sp_get_all_casos(f=None):
    sql="SELECT * FROM v_casos WHERE 1=1"; p=[]
    if f:
        if f.get("estado"): sql+=" AND estado=?"; p.append(f["estado"])
        if f.get("rama_derecho"): sql+=" AND rama_derecho=?"; p.append(f["rama_derecho"])
        if f.get("fecha_desde"): sql+=" AND fecha_apertura>=?"; p.append(f["fecha_desde"])
        if f.get("fecha_hasta"): sql+=" AND fecha_apertura<=?"; p.append(f["fecha_hasta"])
    with get_conn() as c: return c.execute(sql, p).fetchall()
def sp_get_caso(num):
    with get_conn() as c: return c.execute("SELECT * FROM casos WHERE numero_caso=?", (num,)).fetchone()

def sp_insert_audiencia(d):
    with get_conn() as c: c.execute("INSERT INTO audiencias(codigo,tipo,id_caso,fecha_hora,duracion_estimada,lugar,participantes_int,participantes_ext,proposito,resultado_esperado,resultado_real) VALUES(:codigo,:tipo,:id_caso,:fecha_hora,:duracion_estimada,:lugar,:participantes_int,:participantes_ext,:proposito,:resultado_esperado,:resultado_real)", d)
def sp_update_audiencia(d):
    with get_conn() as c: c.execute("UPDATE audiencias SET tipo=:tipo,id_caso=:id_caso,fecha_hora=:fecha_hora,duracion_estimada=:duracion_estimada,lugar=:lugar,participantes_int=:participantes_int,participantes_ext=:participantes_ext,proposito=:proposito,resultado_esperado=:resultado_esperado,resultado_real=:resultado_real WHERE codigo=:codigo", d)
def sp_delete_audiencia(cod):
    with get_conn() as c: c.execute("DELETE FROM audiencias WHERE codigo=?", (cod,))
def sp_get_all_audiencias(f=None):
    sql="SELECT * FROM v_audiencias WHERE 1=1"; p=[]
    if f:
        if f.get("tipo"): sql+=" AND tipo=?"; p.append(f["tipo"])
        if f.get("fecha_desde"): sql+=" AND fecha_hora>=?"; p.append(f["fecha_desde"])
        if f.get("fecha_hasta"): sql+=" AND fecha_hora<=?"; p.append(f["fecha_hasta"])
    with get_conn() as c: return c.execute(sql, p).fetchall()
def sp_get_audiencia(cod):
    with get_conn() as c: return c.execute("SELECT * FROM audiencias WHERE codigo=?", (cod,)).fetchone()

def get_clientes_combo():
    with get_conn() as c:
        rows = c.execute("SELECT id_cliente, nombre FROM clientes ORDER BY nombre").fetchall()
    return {r["nombre"]: r["id_cliente"] for r in rows}
def get_abogados_combo():
    with get_conn() as c:
        rows = c.execute("SELECT id_abogado, nombres||' '||apellidos AS n FROM abogados ORDER BY apellidos").fetchall()
    return {r["n"]: r["id_abogado"] for r in rows}
def get_casos_combo():
    with get_conn() as c:
        rows = c.execute("SELECT id_caso, numero_caso FROM casos ORDER BY numero_caso").fetchall()
    return {r["numero_caso"]: r["id_caso"] for r in rows}

# ─── VALIDACIONES ─────────────────────────────────────────────

def validar_email(v):
    return bool(re.match(r'^[\w\.-]+@[\w\.-]+\.\w{2,}$', v))
def validar_numerico(v):
    try: float(v); return True
    except: return False
def solo_numeros(e, widget):
    v = widget.get()
    limpio = re.sub(r'[^\d.]', '', v)
    if v != limpio:
        widget.delete(0, tk.END); widget.insert(0, limpio)
def validar_cliente(d):
    err = []
    if len(d.get("codigo","").strip()) < 3: err.append("• Codigo: minimo 3 caracteres")
    if len(d.get("nombre","").strip()) < 3: err.append("• Nombre: minimo 3 caracteres")
    if len(d.get("documento","").strip()) < 5: err.append("• Documento: minimo 5 caracteres")
    if d.get("correo") and not validar_email(d["correo"]): err.append("• Correo: formato invalido")
    return err
def validar_abogado(d):
    err = []
    if len(d.get("num_colegiatura","").strip()) < 3: err.append("• N Colegiatura: minimo 3 caracteres")
    if len(d.get("nombres","").strip()) < 2: err.append("• Nombres: minimo 2 caracteres")
    if len(d.get("apellidos","").strip()) < 2: err.append("• Apellidos: minimo 2 caracteres")
    if d.get("anios_exp") and not validar_numerico(d["anios_exp"]): err.append("• Anos experiencia: debe ser numero")
    if d.get("tarifa_hora") and not validar_numerico(d["tarifa_hora"]): err.append("• Tarifa: debe ser numero")
    return err
def validar_caso(d):
    err = []
    if len(d.get("numero_caso","").strip()) < 3: err.append("• Numero de caso: minimo 3 caracteres")
    if len(d.get("titulo","").strip()) < 5: err.append("• Titulo: minimo 5 caracteres")
    return err
def validar_audiencia(d):
    err = []
    if len(d.get("codigo","").strip()) < 3: err.append("• Codigo: minimo 3 caracteres")
    if len(d.get("lugar","").strip()) < 3: err.append("• Lugar: minimo 3 caracteres")
    return err
def mostrar_errores(errores):
    if errores:
        messagebox.showerror("Errores de validacion", "\n".join(errores)); return True
    return False

# ─── EXPORTACIÓN EXCEL ────────────────────────────────────────

def exportar_excel(titulo, encabezados, filas, filtro_info=""):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from datetime import datetime as dt
        ruta = filedialog.asksaveasfilename(defaultextension=".xlsx",
            filetypes=[("Excel","*.xlsx")],
            initialfile=f"{titulo.replace(' ','_')}_{dt.now().strftime('%Y%m%d')}.xlsx")
        if not ruta: return
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = titulo[:30]
        fn = PatternFill("solid",fgColor="1E2761"); fg = PatternFill("solid",fgColor="C9A84C")
        fp = PatternFill("solid",fgColor="F2F4F8")
        borde = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
        centro = Alignment(horizontal="center",vertical="center")
        ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=len(encabezados))
        c = ws.cell(row=1,column=1,value=f"JUSTICIA & ASOCIADOS - {titulo.upper()}")
        c.font=Font(name="Calibri",size=13,bold=True,color="FFFFFF"); c.fill=fn; c.alignment=centro
        ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=len(encabezados))
        info = f"Exportado: {dt.now().strftime('%d/%m/%Y %H:%M')}"
        if filtro_info: info += f"  |  Filtro: {filtro_info}"
        ws.cell(row=2,column=1,value=info).alignment=centro
        for col,enc in enumerate(encabezados,1):
            c=ws.cell(row=3,column=col,value=enc)
            c.font=Font(name="Calibri",size=10,bold=True,color="FFFFFF"); c.fill=fg; c.alignment=centro; c.border=borde
        for i,fila in enumerate(filas):
            for col,val in enumerate(fila,1):
                c=ws.cell(row=i+4,column=col,value=val); c.border=borde; c.alignment=Alignment(vertical="center")
                if i%2==0: c.fill=fp
        for col in ws.columns:
            mx=max((len(str(c.value or "")) for c in col),default=10)
            ws.column_dimensions[col[0].column_letter].width=min(mx+4,40)
        wb.save(ruta); messagebox.showinfo("Exito",f"Excel guardado:\n{ruta}")
    except Exception as e: messagebox.showerror("Error Excel",str(e))

# ─── EXPORTACIÓN PDF ──────────────────────────────────────────

def exportar_pdf(titulo, encabezados, filas, filtro_info=""):
    try:
        from reportlab.lib.pagesizes import landscape, A4
        from reportlab.lib import colors
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from datetime import datetime as dt
        ruta = filedialog.asksaveasfilename(defaultextension=".pdf",
            filetypes=[("PDF","*.pdf")],
            initialfile=f"{titulo.replace(' ','_')}_{dt.now().strftime('%Y%m%d')}.pdf")
        if not ruta: return
        doc = SimpleDocTemplate(ruta,pagesize=landscape(A4),leftMargin=1.5*cm,rightMargin=1.5*cm,topMargin=2*cm,bottomMargin=2*cm)
        elems=[]; navy=colors.HexColor("#1E2761"); gold=colors.HexColor("#C9A84C")
        elems.append(Paragraph("JUSTICIA & ASOCIADOS",ParagraphStyle("t",fontSize=16,fontName="Helvetica-Bold",textColor=navy,alignment=1,spaceAfter=4)))
        elems.append(Paragraph(titulo.upper(),ParagraphStyle("s",fontSize=12,fontName="Helvetica-Bold",textColor=gold,alignment=1,spaceAfter=4)))
        info=f"Exportado: {dt.now().strftime('%d/%m/%Y %H:%M')}"
        if filtro_info: info+=f"   |   Filtro: {filtro_info}"
        elems.append(Paragraph(info,ParagraphStyle("i",fontSize=9,fontName="Helvetica",textColor=colors.grey,alignment=1,spaceAfter=10)))
        elems.append(Spacer(1,0.3*cm))
        ancho=(landscape(A4)[0]-3*cm)/len(encabezados)
        data=[encabezados]+[list(f) for f in filas]
        tabla=Table(data,colWidths=[ancho]*len(encabezados),repeatRows=1)
        estilo=TableStyle([("BACKGROUND",(0,0),(-1,0),navy),("TEXTCOLOR",(0,0),(-1,0),gold),
            ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),("FONTSIZE",(0,0),(-1,0),9),
            ("ALIGN",(0,0),(-1,-1),"CENTER"),("VALIGN",(0,0),(-1,-1),"MIDDLE"),
            ("FONTNAME",(0,1),(-1,-1),"Helvetica"),("FONTSIZE",(0,1),(-1,-1),8),
            ("GRID",(0,0),(-1,-1),0.5,colors.HexColor("#CCCCCC")),
            ("TOPPADDING",(0,0),(-1,-1),5),("BOTTOMPADDING",(0,0),(-1,-1),5)])
        for i in range(1,len(data)):
            if i%2==0: estilo.add("BACKGROUND",(0,i),(-1,i),colors.HexColor("#F2F4F8"))
        tabla.setStyle(estilo); elems.append(tabla); elems.append(Spacer(1,0.4*cm))
        elems.append(Paragraph(f"Total registros: {len(filas)}",ParagraphStyle("p",fontSize=8,fontName="Helvetica",textColor=colors.grey,alignment=2)))
        doc.build(elems); messagebox.showinfo("Exito",f"PDF guardado:\n{ruta}")
    except Exception as e: messagebox.showerror("Error PDF",str(e))

# ─── IMÁGENES ─────────────────────────────────────────────────

def cargar_imagen(label_widget):
    try:
        from PIL import Image, ImageTk
        ruta = filedialog.askopenfilename(title="Seleccionar imagen",filetypes=[("Imagenes","*.jpg *.jpeg *.png *.gif")])
        if not ruta: return None
        ext = os.path.splitext(ruta)[1].lower()
        if ext not in (".jpg",".jpeg",".png",".gif"):
            messagebox.showerror("Formato invalido","Solo JPG, PNG y GIF."); return None
        if os.path.getsize(ruta)>5*1024*1024:
            messagebox.showerror("Archivo grande","La imagen no debe superar 5 MB."); return None
        img=Image.open(ruta).convert("RGB"); img.thumbnail((150,150),Image.LANCZOS)
        foto=ImageTk.PhotoImage(img); label_widget.config(image=foto,text=""); label_widget.image=foto
        buf=io.BytesIO(); img.save(buf,format="PNG"); return buf.getvalue()
    except Exception as e: messagebox.showerror("Error imagen",str(e)); return None

def mostrar_img_bytes(label_widget, datos):
    try:
        from PIL import Image, ImageTk
        if not datos: label_widget.config(image="",text="Sin foto"); return
        img=Image.open(io.BytesIO(datos)); img.thumbnail((150,150),Image.LANCZOS)
        foto=ImageTk.PhotoImage(img); label_widget.config(image=foto,text=""); label_widget.image=foto
    except: label_widget.config(image="",text="Sin foto")

def crear_favicon(root):
    try:
        from PIL import Image, ImageDraw, ImageTk
        img=Image.new("RGBA",(64,64),(30,39,97,255)); d=ImageDraw.Draw(img)
        d.rectangle([28,10,36,50],fill=(201,168,76,255)); d.ellipse([18,10,46,24],fill=(201,168,76,255))
        d.ellipse([12,24,30,36],fill=(201,168,76,255)); d.ellipse([34,24,52,36],fill=(201,168,76,255))
        d.rectangle([20,48,44,52],fill=(201,168,76,255))
        ico=ImageTk.PhotoImage(img); root.iconphoto(True,ico); root._fav=ico
    except: pass

# ─── TEMAS ────────────────────────────────────────────────────

TEMAS = {
    "Claro":{"bg":"#F0F2F8","entry_bg":"#FFFFFF","entry_fg":"#2D3748","fg":"#2D3748","frame_bg":"#FFFFFF"},
    "Oscuro":{"bg":"#1A1A2E","entry_bg":"#0F3460","entry_fg":"#E0E0E0","fg":"#E0E0E0","frame_bg":"#16213E"},
}
def aplicar_tema(root, nombre):
    t=TEMAS[nombre]; root.config(bg=t["bg"])
    def _rec(w):
        try:
            cls=w.winfo_class()
            if cls in ("Frame","Labelframe"): w.config(bg=t["frame_bg"])
            elif cls=="Label": w.config(bg=t["bg"],fg=t["fg"])
            elif cls=="Entry": w.config(bg=t["entry_bg"],fg=t["entry_fg"],insertbackground=t["fg"])
            elif cls=="Text": w.config(bg=t["entry_bg"],fg=t["entry_fg"],insertbackground=t["fg"])
        except: pass
        for ch in w.winfo_children(): _rec(ch)
    _rec(root)

# ─── HELPERS UI ───────────────────────────────────────────────

def make_scroll(parent):
    c=tk.Canvas(parent,highlightthickness=0); sb=ttk.Scrollbar(parent,orient="vertical",command=c.yview)
    inner=tk.Frame(c)
    inner.bind("<Configure>",lambda e: c.configure(scrollregion=c.bbox("all")))
    c.create_window((0,0),window=inner,anchor="nw"); c.configure(yscrollcommand=sb.set)
    c.pack(side="left",fill="both",expand=True); sb.pack(side="right",fill="y")
    return inner

def lbl(p,t,r,c=0):
    tk.Label(p,text=t,font=("Calibri",11),fg="#1E2761",anchor="w").grid(row=r,column=c,sticky="w",padx=(0,10),pady=6)

def ent(p,r,c=1,w=28):
    e=tk.Entry(p,width=w,font=("Calibri",11),relief="solid",bd=1,bg="#FFFFFF",fg="#2D3748")
    e.grid(row=r,column=c,sticky="w",pady=6); return e

def combo(p,var,vals,r,c=1):
    cb=ttk.Combobox(p,textvariable=var,values=vals,width=26,state="readonly")
    cb.grid(row=r,column=c,sticky="w",pady=6); return cb

def make_tree(parent,cols,height=6):
    f=tk.Frame(parent); f.pack(fill="x",padx=40,pady=5)
    tv=ttk.Treeview(f,columns=cols,show="headings",height=height)
    for col in cols: tv.heading(col,text=col); tv.column(col,width=130,anchor="center")
    sb=ttk.Scrollbar(f,orient="horizontal",command=tv.xview)
    tv.configure(xscrollcommand=sb.set); tv.pack(fill="x"); sb.pack(fill="x")
    return tv

def hacer_botones(parent,cmds):
    colores={"Guardar":("#4CAF50","Guardar"),"Actualizar":("#2196F3","Actualizar"),
             "Eliminar":("#f44336","Eliminar"),"Limpiar":("#FF9800","Limpiar"),
             "Excel":("#7B1FA2","Exportar Excel"),"PDF":("#D32F2F","Exportar PDF")}
    f=tk.Frame(parent); f.pack(pady=10)
    for nombre,fn in cmds.items():
        color,texto=colores.get(nombre,("#607D8B",nombre))
        tk.Button(f,text=texto,font=("Calibri",10,"bold"),bg=color,fg="white",
                  padx=10,pady=4,relief="flat",cursor="hand2",command=fn).pack(side=tk.LEFT,padx=4)

def hacer_filtro(parent,widgets_fn):
    f=tk.LabelFrame(parent,text=" Filtros de Exportacion ",font=("Calibri",10,"bold"),fg="#1E2761",padx=10,pady=5)
    f.pack(fill="x",padx=40,pady=(5,0)); widgets_fn(f)

# ════════════════════════════════════════════════════════════════
#  VENTANA PRINCIPAL
# ════════════════════════════════════════════════════════════════

init_db()
root = tk.Tk()
root.geometry("1100x720")
root.title("Justicia & Asociados - LawFirm Management")
root.minsize(900,600)
crear_favicon(root)

menubar=tk.Menu(root)
m=tk.Menu(menubar,tearoff=0)
m.add_command(label="Tema Claro", command=lambda: aplicar_tema(root,"Claro"))
m.add_command(label="Tema Oscuro",command=lambda: aplicar_tema(root,"Oscuro"))
menubar.add_cascade(label="Tema",menu=m)
root.config(menu=menubar)

notebook=ttk.Notebook(root)
tab1=ttk.Frame(notebook); tab2=ttk.Frame(notebook)
tab3=ttk.Frame(notebook); tab4=ttk.Frame(notebook)
notebook.add(tab1,text="Clientes")
notebook.add(tab2,text="Abogados")
notebook.add(tab3,text="Casos Legales")
notebook.add(tab4,text="Audiencias y Citas")
notebook.pack(expand=True,fill="both")

# ════════════════════════════════════════════════════════════════
#  MÓDULO 1 — CLIENTES
# ════════════════════════════════════════════════════════════════

foto_cli=[None]
inner1=make_scroll(tab1)
tk.Label(inner1,text="GESTION DE CLIENTES",font=("Georgia",15,"bold"),fg="#1E2761").pack(pady=(15,5))

pane1=tk.Frame(inner1); pane1.pack(fill="x",padx=40,pady=5)
f1=tk.Frame(pane1); f1.grid(row=0,column=0,sticky="nw")
fp1=tk.Frame(pane1); fp1.grid(row=0,column=1,sticky="n",padx=(30,0))

lbl(f1,"Codigo Cliente: *",1);        c_cod=ent(f1,1)
lbl(f1,"Tipo Cliente: *",2);          c_tipo=tk.StringVar(); combo(f1,c_tipo,["Persona Natural","Empresa"],2)
lbl(f1,"Nombre / Razon Social: *",3); c_nom=ent(f1,3)
lbl(f1,"Documento / RUC: *",4);       c_doc=ent(f1,4)
lbl(f1,"Direccion:",5);               c_dir=ent(f1,5)
lbl(f1,"Telefono:",6);                c_tel=ent(f1,6)
lbl(f1,"Correo Electronico:",7);      c_cor=ent(f1,7)
lbl(f1,"Fecha Primer Contacto:",8)
c_fec=DateEntry(f1,width=26,date_pattern="yyyy-mm-dd",font=("Calibri",11)); c_fec.grid(row=8,column=1,sticky="w",pady=6)
lbl(f1,"Referido Por:",9);            c_ref=ent(f1,9)
lbl(f1,"Clasificacion Interna:",10);  c_cla=tk.StringVar()
combo(f1,c_cla,["Cliente VIP","Estandar","Corporativo","Pro Bono"],10)

tk.Label(fp1,text="Foto del Cliente",font=("Calibri",10,"bold"),fg="#1E2761").pack()
lbl_fc=tk.Label(fp1,text="Sin foto",width=20,height=8,relief="solid",bd=1,bg="#F0F2F8"); lbl_fc.pack(pady=5)
tk.Button(fp1,text="Seleccionar foto",font=("Calibri",9),bg="#1E2761",fg="white",
          command=lambda:[foto_cli.__setitem__(0,cargar_imagen(lbl_fc))]).pack(fill="x",pady=2)
tk.Button(fp1,text="Quitar foto",font=("Calibri",9),bg="#607D8B",fg="white",
          command=lambda:[foto_cli.__setitem__(0,None),lbl_fc.config(image="",text="Sin foto")]).pack(fill="x")

# --- Variables de filtro 1 ---
f1_tipo=tk.StringVar(); f1_cla2=tk.StringVar()
f1_des=[None]; f1_has=[None]

# --- Funciones CRUD 1 (definidas ANTES de usarlas) ---
def get_d1():
    return {"codigo":c_cod.get().strip(),"tipo":c_tipo.get(),"nombre":c_nom.get().strip(),
            "documento":c_doc.get().strip(),"direccion":c_dir.get().strip(),
            "telefono":c_tel.get().strip(),"correo":c_cor.get().strip(),
            "fecha_contacto":c_fec.get(),"referido_por":c_ref.get().strip(),
            "clasificacion":c_cla.get(),"foto":foto_cli[0]}

def refresh1():
    tree1.delete(*tree1.get_children())
    for r in sp_get_all_clientes({"tipo":f1_tipo.get() or None,
        "fecha_desde":f1_des[0].get() if f1_des[0] else None,
        "fecha_hasta":f1_has[0].get() if f1_has[0] else None,
        "clasificacion":f1_cla2.get() or None}):
        tree1.insert("","end",values=(r["codigo"],r["tipo"],r["nombre"],
            r["documento"],r["telefono"] or "",r["correo"] or "",r["clasificacion"] or ""))

def on_sel1(e):
    s=tree1.selection()
    if not s: return
    r=sp_get_cliente(str(tree1.item(s[0])["values"][0]))
    if not r: return
    for w,k in [(c_cod,"codigo"),(c_nom,"nombre"),(c_doc,"documento"),
                (c_dir,"direccion"),(c_tel,"telefono"),(c_cor,"correo"),(c_ref,"referido_por")]:
        w.delete(0,tk.END); w.insert(0,r[k] or "")
    c_tipo.set(r["tipo"] or ""); c_cla.set(r["clasificacion"] or "")
    mostrar_img_bytes(lbl_fc,r["foto"]); foto_cli[0]=r["foto"]

def save1():
    d=get_d1(); e=validar_cliente(d)
    if mostrar_errores(e): return
    try: sp_insert_cliente(d); messagebox.showinfo("Guardado","Cliente guardado correctamente."); clear1(); refresh1()
    except Exception as ex: messagebox.showerror("Error",str(ex))

def update1():
    if not messagebox.askyesno("Confirmar","Actualizar este cliente?"): return
    d=get_d1(); e=validar_cliente(d)
    if mostrar_errores(e): return
    try: sp_update_cliente(d); messagebox.showinfo("Actualizado","Cliente actualizado correctamente."); refresh1()
    except Exception as ex: messagebox.showerror("Error",str(ex))

def delete1():
    cod=c_cod.get().strip()
    if not cod: messagebox.showwarning("Aviso","Selecciona un cliente del listado."); return
    if not messagebox.askyesno("Eliminar",f"Eliminar '{cod}'?\nEsta accion no se puede deshacer."): return
    try: sp_delete_cliente(cod); messagebox.showinfo("Eliminado","Cliente eliminado."); clear1(); refresh1()
    except Exception as ex: messagebox.showerror("Error",str(ex))

def clear1():
    for w in [c_cod,c_nom,c_doc,c_dir,c_tel,c_cor,c_ref]: w.delete(0,tk.END)
    c_tipo.set(""); c_cla.set(""); foto_cli[0]=None; lbl_fc.config(image="",text="Sin foto")

def excel1():
    rows=sp_get_all_clientes({"tipo":f1_tipo.get() or None,
        "fecha_desde":f1_des[0].get() if f1_des[0] else None,
        "fecha_hasta":f1_has[0].get() if f1_has[0] else None,
        "clasificacion":f1_cla2.get() or None})
    exportar_excel("Clientes",["Codigo","Tipo","Nombre","Documento","Telefono","Correo","Clasificacion"],
        [(r["codigo"],r["tipo"],r["nombre"],r["documento"],r["telefono"] or "",r["correo"] or "",r["clasificacion"] or "") for r in rows],
        f"Tipo:{f1_tipo.get() or 'Todos'}")

def pdf1():
    rows=sp_get_all_clientes({"tipo":f1_tipo.get() or None,
        "fecha_desde":f1_des[0].get() if f1_des[0] else None,
        "fecha_hasta":f1_has[0].get() if f1_has[0] else None,
        "clasificacion":f1_cla2.get() or None})
    exportar_pdf("Clientes",["Codigo","Tipo","Nombre","Documento","Telefono","Correo","Clasificacion"],
        [(r["codigo"],r["tipo"],r["nombre"],r["documento"],r["telefono"] or "",r["correo"] or "",r["clasificacion"] or "") for r in rows],
        f"Tipo:{f1_tipo.get() or 'Todos'}")

# --- Filtros y tree (DESPUÉS de definir refresh1) ---
def build_filtros1(f):
    tk.Label(f,text="Tipo:",font=("Calibri",10)).grid(row=0,column=0,padx=5)
    ttk.Combobox(f,textvariable=f1_tipo,values=["","Persona Natural","Empresa"],width=14,state="readonly").grid(row=0,column=1,padx=5)
    tk.Label(f,text="Desde:",font=("Calibri",10)).grid(row=0,column=2,padx=5)
    f1_des[0]=DateEntry(f,width=12,date_pattern="yyyy-mm-dd"); f1_des[0].grid(row=0,column=3,padx=5)
    tk.Label(f,text="Hasta:",font=("Calibri",10)).grid(row=0,column=4,padx=5)
    f1_has[0]=DateEntry(f,width=12,date_pattern="yyyy-mm-dd"); f1_has[0].grid(row=0,column=5,padx=5)
    tk.Label(f,text="Clasificacion:",font=("Calibri",10)).grid(row=0,column=6,padx=5)
    ttk.Combobox(f,textvariable=f1_cla2,values=["","Cliente VIP","Estandar","Corporativo","Pro Bono"],width=12,state="readonly").grid(row=0,column=7,padx=5)
    tk.Button(f,text="Filtrar",font=("Calibri",9,"bold"),bg="#1E2761",fg="white",command=refresh1).grid(row=0,column=8,padx=10)

hacer_filtro(inner1,build_filtros1)
tree1=make_tree(inner1,("Codigo","Tipo","Nombre","Documento","Telefono","Correo","Clasificacion"))
tree1.bind("<<TreeviewSelect>>",on_sel1)
hacer_botones(inner1,{"Guardar":save1,"Actualizar":update1,"Eliminar":delete1,"Limpiar":clear1,"Excel":excel1,"PDF":pdf1})
refresh1()

# ════════════════════════════════════════════════════════════════
#  MÓDULO 2 — ABOGADOS
# ════════════════════════════════════════════════════════════════

foto_abo=[None]
inner2=make_scroll(tab2)
tk.Label(inner2,text="GESTION DE ABOGADOS",font=("Georgia",15,"bold"),fg="#1E2761").pack(pady=(15,5))

pane2=tk.Frame(inner2); pane2.pack(fill="x",padx=40,pady=5)
f2=tk.Frame(pane2); f2.grid(row=0,column=0,sticky="nw")
fp2=tk.Frame(pane2); fp2.grid(row=0,column=1,sticky="n",padx=(30,0))

lbl(f2,"N Colegiatura: *",1);    a_col=ent(f2,1)
lbl(f2,"Nombres: *",2);          a_nom=ent(f2,2)
lbl(f2,"Apellidos: *",3);        a_ape=ent(f2,3)
lbl(f2,"Especialidad:",4);       a_esp=tk.StringVar()
combo(f2,a_esp,["Civil","Penal","Laboral","Tributario","Mercantil"],4)
lbl(f2,"Anos de Experiencia:",5);a_ani=ent(f2,5)
a_ani.bind("<KeyRelease>",lambda e:solo_numeros(e,a_ani))
lbl(f2,"Formacion Academica:",6);a_for=ent(f2,6)
lbl(f2,"Idiomas:",7);            a_idi=ent(f2,7)
lbl(f2,"Tarifa por Hora ($):",8);a_tar=ent(f2,8)
a_tar.bind("<KeyRelease>",lambda e:solo_numeros(e,a_tar))
lbl(f2,"Disponibilidad:",9);     a_dis=tk.StringVar(value="Disponible")
combo(f2,a_dis,["Disponible","Ocupado","De vacaciones"],9)

tk.Label(fp2,text="Foto del Abogado",font=("Calibri",10,"bold"),fg="#1E2761").pack()
lbl_fa=tk.Label(fp2,text="Sin foto",width=20,height=8,relief="solid",bd=1,bg="#F0F2F8"); lbl_fa.pack(pady=5)
tk.Button(fp2,text="Seleccionar foto",font=("Calibri",9),bg="#1E2761",fg="white",
          command=lambda:[foto_abo.__setitem__(0,cargar_imagen(lbl_fa))]).pack(fill="x",pady=2)
tk.Button(fp2,text="Quitar foto",font=("Calibri",9),bg="#607D8B",fg="white",
          command=lambda:[foto_abo.__setitem__(0,None),lbl_fa.config(image="",text="Sin foto")]).pack(fill="x")

f2_esp=tk.StringVar(); f2_dis=tk.StringVar()

def get_d2():
    return {"num_colegiatura":a_col.get().strip(),"nombres":a_nom.get().strip(),
            "apellidos":a_ape.get().strip(),"especialidad":a_esp.get(),
            "anios_exp":a_ani.get().strip() or None,"formacion":a_for.get().strip(),
            "idiomas":a_idi.get().strip(),"tarifa_hora":a_tar.get().strip() or None,
            "disponibilidad":a_dis.get(),"foto":foto_abo[0]}

def refresh2():
    tree2.delete(*tree2.get_children())
    for r in sp_get_all_abogados({"especialidad":f2_esp.get() or None,"disponibilidad":f2_dis.get() or None}):
        tree2.insert("","end",values=(r["num_colegiatura"],f"{r['nombres']} {r['apellidos']}",
            r["especialidad"],r["anios_exp"] or "",
            f"${r['tarifa_hora']:,.0f}" if r["tarifa_hora"] else "",r["disponibilidad"]))

def on_sel2(e):
    s=tree2.selection()
    if not s: return
    r=sp_get_abogado(str(tree2.item(s[0])["values"][0]))
    if not r: return
    for w,k in [(a_col,"num_colegiatura"),(a_nom,"nombres"),(a_ape,"apellidos"),
                (a_ani,"anios_exp"),(a_for,"formacion"),(a_idi,"idiomas"),(a_tar,"tarifa_hora")]:
        w.delete(0,tk.END); w.insert(0,str(r[k] or ""))
    a_esp.set(r["especialidad"] or ""); a_dis.set(r["disponibilidad"])
    mostrar_img_bytes(lbl_fa,r["foto"]); foto_abo[0]=r["foto"]

def save2():
    d=get_d2(); e=validar_abogado(d)
    if mostrar_errores(e): return
    try: sp_insert_abogado(d); messagebox.showinfo("Guardado","Abogado guardado correctamente."); clear2(); refresh2()
    except Exception as ex: messagebox.showerror("Error",str(ex))

def update2():
    if not messagebox.askyesno("Confirmar","Actualizar este abogado?"): return
    d=get_d2(); e=validar_abogado(d)
    if mostrar_errores(e): return
    try: sp_update_abogado(d); messagebox.showinfo("Actualizado","Abogado actualizado correctamente."); refresh2()
    except Exception as ex: messagebox.showerror("Error",str(ex))

def delete2():
    col=a_col.get().strip()
    if not col: messagebox.showwarning("Aviso","Selecciona un abogado del listado."); return
    if not messagebox.askyesno("Eliminar",f"Eliminar '{col}'?"): return
    try: sp_delete_abogado(col); messagebox.showinfo("Eliminado","Abogado eliminado."); clear2(); refresh2()
    except Exception as ex: messagebox.showerror("Error",str(ex))

def clear2():
    for w in [a_col,a_nom,a_ape,a_ani,a_for,a_idi,a_tar]: w.delete(0,tk.END)
    a_esp.set(""); a_dis.set("Disponible"); foto_abo[0]=None; lbl_fa.config(image="",text="Sin foto")

def excel2():
    rows=sp_get_all_abogados({"especialidad":f2_esp.get() or None,"disponibilidad":f2_dis.get() or None})
    exportar_excel("Abogados",["Colegiatura","Nombres","Apellidos","Especialidad","Anos","Tarifa","Disponibilidad"],
        [(r["num_colegiatura"],r["nombres"],r["apellidos"],r["especialidad"],
          r["anios_exp"] or "",r["tarifa_hora"] or "",r["disponibilidad"]) for r in rows],
        f"Especialidad:{f2_esp.get() or 'Todas'}")

def pdf2():
    rows=sp_get_all_abogados({"especialidad":f2_esp.get() or None,"disponibilidad":f2_dis.get() or None})
    exportar_pdf("Abogados",["Colegiatura","Nombre Completo","Especialidad","Tarifa/Hora","Disponibilidad"],
        [(r["num_colegiatura"],f"{r['nombres']} {r['apellidos']}",r["especialidad"],
          r["tarifa_hora"] or "",r["disponibilidad"]) for r in rows],
        f"Especialidad:{f2_esp.get() or 'Todas'}")

def build_filtros2(f):
    tk.Label(f,text="Especialidad:",font=("Calibri",10)).grid(row=0,column=0,padx=5)
    ttk.Combobox(f,textvariable=f2_esp,values=["","Civil","Penal","Laboral","Tributario","Mercantil"],width=14,state="readonly").grid(row=0,column=1,padx=5)
    tk.Label(f,text="Disponibilidad:",font=("Calibri",10)).grid(row=0,column=2,padx=5)
    ttk.Combobox(f,textvariable=f2_dis,values=["","Disponible","Ocupado","De vacaciones"],width=14,state="readonly").grid(row=0,column=3,padx=5)
    tk.Button(f,text="Filtrar",font=("Calibri",9,"bold"),bg="#1E2761",fg="white",command=refresh2).grid(row=0,column=4,padx=10)

hacer_filtro(inner2,build_filtros2)
tree2=make_tree(inner2,("Colegiatura","Nombre Completo","Especialidad","Anos Exp.","Tarifa/Hora","Disponibilidad"))
tree2.bind("<<TreeviewSelect>>",on_sel2)
hacer_botones(inner2,{"Guardar":save2,"Actualizar":update2,"Eliminar":delete2,"Limpiar":clear2,"Excel":excel2,"PDF":pdf2})
refresh2()

# ════════════════════════════════════════════════════════════════
#  MÓDULO 3 — CASOS LEGALES
# ════════════════════════════════════════════════════════════════

inner3=make_scroll(tab3)
tk.Label(inner3,text="GESTION DE CASOS LEGALES",font=("Georgia",15,"bold"),fg="#1E2761").pack(pady=(15,5))
f3=tk.Frame(inner3); f3.pack(fill="x",padx=60,pady=5)

lbl(f3,"Numero de Caso: *",1);      ca_num=ent(f3,1)
lbl(f3,"Titulo Descriptivo: *",2);  ca_tit=ent(f3,2,w=40)
lbl(f3,"Tipo de Caso:",3);          ca_tip=ent(f3,3)
lbl(f3,"Rama del Derecho:",4);      ca_ram=tk.StringVar()
combo(f3,ca_ram,["Civil","Penal","Laboral","Tributario","Mercantil","Administrativo"],4)
lbl(f3,"Fecha de Apertura:",5)
ca_fec=DateEntry(f3,width=26,date_pattern="yyyy-mm-dd",font=("Calibri",11)); ca_fec.grid(row=5,column=1,sticky="w",pady=6)
lbl(f3,"Cliente:",6);               ca_cli_v=tk.StringVar()
ca_cli_cb=ttk.Combobox(f3,textvariable=ca_cli_v,width=26,state="readonly"); ca_cli_cb.grid(row=6,column=1,sticky="w",pady=6)
lbl(f3,"Contraparte:",7);           ca_con=ent(f3,7)
lbl(f3,"Juzgado / Entidad:",8);     ca_juz=ent(f3,8)
lbl(f3,"N Expediente Externo:",9);  ca_exp=ent(f3,9)
lbl(f3,"Abogado Principal:",10);    ca_abo_v=tk.StringVar()
ca_abo_cb=ttk.Combobox(f3,textvariable=ca_abo_v,width=26,state="readonly"); ca_abo_cb.grid(row=10,column=1,sticky="w",pady=6)
lbl(f3,"Estado Actual:",11);        ca_est=tk.StringVar(value="Abierto")
combo(f3,ca_est,["Abierto","En proceso","Suspendido","Cerrado","Ganado","Perdido"],11)
lbl(f3,"Fecha Est. Conclusion:",12)
ca_con2=DateEntry(f3,width=26,date_pattern="yyyy-mm-dd",font=("Calibri",11)); ca_con2.grid(row=12,column=1,sticky="w",pady=6)

f3_est=tk.StringVar(); f3_ram=tk.StringVar()
f3_des3=[None]; f3_has3=[None]

def refresh_combos3():
    cli=get_clientes_combo(); abo=get_abogados_combo()
    ca_cli_cb["values"]=list(cli.keys()); ca_cli_cb._cli=cli
    ca_abo_cb["values"]=list(abo.keys()); ca_abo_cb._abo=abo

def get_d3():
    cli=getattr(ca_cli_cb,"_cli",{}); abo=getattr(ca_abo_cb,"_abo",{})
    return {"numero_caso":ca_num.get().strip(),"titulo":ca_tit.get().strip(),
            "tipo_caso":ca_tip.get().strip(),"rama_derecho":ca_ram.get(),
            "fecha_apertura":ca_fec.get(),"id_cliente":cli.get(ca_cli_v.get()),
            "contraparte":ca_con.get().strip(),"juzgado":ca_juz.get().strip(),
            "num_expediente":ca_exp.get().strip(),"id_abogado_principal":abo.get(ca_abo_v.get()),
            "estado":ca_est.get(),"fecha_conclusion":ca_con2.get()}

def refresh3():
    tree3.delete(*tree3.get_children())
    for r in sp_get_all_casos({"estado":f3_est.get() or None,"rama_derecho":f3_ram.get() or None,
        "fecha_desde":f3_des3[0].get() if f3_des3[0] else None,
        "fecha_hasta":f3_has3[0].get() if f3_has3[0] else None}):
        tree3.insert("","end",values=(r["numero_caso"],r["titulo"][:30],r["rama_derecho"] or "",
            r["cliente"] or "",r["abogado_principal"] or "",r["estado"],r["fecha_apertura"] or ""))

def on_sel3(e):
    s=tree3.selection()
    if not s: return
    r=sp_get_caso(str(tree3.item(s[0])["values"][0]))
    if not r: return
    for w,k in [(ca_num,"numero_caso"),(ca_tit,"titulo"),(ca_tip,"tipo_caso"),
                (ca_con,"contraparte"),(ca_juz,"juzgado"),(ca_exp,"num_expediente")]:
        w.delete(0,tk.END); w.insert(0,r[k] or "")
    ca_ram.set(r["rama_derecho"] or ""); ca_est.set(r["estado"])

def save3():
    d=get_d3(); e=validar_caso(d)
    if mostrar_errores(e): return
    try: sp_insert_caso(d); messagebox.showinfo("Guardado","Caso guardado correctamente."); clear3(); refresh3(); refresh_combos3()
    except Exception as ex: messagebox.showerror("Error",str(ex))

def update3():
    if not messagebox.askyesno("Confirmar","Actualizar este caso?"): return
    d=get_d3(); e=validar_caso(d)
    if mostrar_errores(e): return
    try: sp_update_caso(d); messagebox.showinfo("Actualizado","Caso actualizado correctamente."); refresh3()
    except Exception as ex: messagebox.showerror("Error",str(ex))

def delete3():
    num=ca_num.get().strip()
    if not num: messagebox.showwarning("Aviso","Selecciona un caso del listado."); return
    if not messagebox.askyesno("Eliminar",f"Eliminar '{num}'? Esta accion es irreversible."): return
    try: sp_delete_caso(num); messagebox.showinfo("Eliminado","Caso eliminado."); clear3(); refresh3()
    except Exception as ex: messagebox.showerror("Error",str(ex))

def clear3():
    for w in [ca_num,ca_tit,ca_tip,ca_con,ca_juz,ca_exp]: w.delete(0,tk.END)
    ca_ram.set(""); ca_est.set("Abierto"); ca_cli_v.set(""); ca_abo_v.set("")

def excel3():
    rows=sp_get_all_casos({"estado":f3_est.get() or None,"rama_derecho":f3_ram.get() or None,
        "fecha_desde":f3_des3[0].get() if f3_des3[0] else None,
        "fecha_hasta":f3_has3[0].get() if f3_has3[0] else None})
    exportar_excel("Casos Legales",["N Caso","Titulo","Rama","Cliente","Abogado","Estado","Apertura","Conclusion"],
        [(r["numero_caso"],r["titulo"],r["rama_derecho"] or "",r["cliente"] or "",
          r["abogado_principal"] or "",r["estado"],r["fecha_apertura"] or "",r["fecha_conclusion"] or "") for r in rows],
        f"Estado:{f3_est.get() or 'Todos'}")

def pdf3():
    rows=sp_get_all_casos({"estado":f3_est.get() or None,"rama_derecho":f3_ram.get() or None,
        "fecha_desde":f3_des3[0].get() if f3_des3[0] else None,
        "fecha_hasta":f3_has3[0].get() if f3_has3[0] else None})
    exportar_pdf("Casos Legales",["N Caso","Titulo","Rama","Estado","Cliente"],
        [(r["numero_caso"],r["titulo"][:35],r["rama_derecho"] or "",r["estado"],r["cliente"] or "") for r in rows],
        f"Estado:{f3_est.get() or 'Todos'}")

def build_filtros3(f):
    tk.Label(f,text="Estado:",font=("Calibri",10)).grid(row=0,column=0,padx=5)
    ttk.Combobox(f,textvariable=f3_est,values=["","Abierto","En proceso","Suspendido","Cerrado","Ganado","Perdido"],width=12,state="readonly").grid(row=0,column=1,padx=5)
    tk.Label(f,text="Rama:",font=("Calibri",10)).grid(row=0,column=2,padx=5)
    ttk.Combobox(f,textvariable=f3_ram,values=["","Civil","Penal","Laboral","Tributario","Mercantil","Administrativo"],width=12,state="readonly").grid(row=0,column=3,padx=5)
    tk.Label(f,text="Desde:",font=("Calibri",10)).grid(row=0,column=4,padx=5)
    f3_des3[0]=DateEntry(f,width=12,date_pattern="yyyy-mm-dd"); f3_des3[0].grid(row=0,column=5,padx=5)
    tk.Label(f,text="Hasta:",font=("Calibri",10)).grid(row=0,column=6,padx=5)
    f3_has3[0]=DateEntry(f,width=12,date_pattern="yyyy-mm-dd"); f3_has3[0].grid(row=0,column=7,padx=5)
    tk.Button(f,text="Filtrar",font=("Calibri",9,"bold"),bg="#1E2761",fg="white",command=refresh3).grid(row=0,column=8,padx=10)

hacer_filtro(inner3,build_filtros3)
tree3=make_tree(inner3,("N Caso","Titulo","Rama","Cliente","Abogado","Estado","Apertura"))
tree3.bind("<<TreeviewSelect>>",on_sel3)
hacer_botones(inner3,{"Guardar":save3,"Actualizar":update3,"Eliminar":delete3,"Limpiar":clear3,"Excel":excel3,"PDF":pdf3})
refresh_combos3()
refresh3()

# ════════════════════════════════════════════════════════════════
#  MÓDULO 4 — AUDIENCIAS Y CITAS
# ════════════════════════════════════════════════════════════════

inner4=make_scroll(tab4)
tk.Label(inner4,text="GESTION DE AUDIENCIAS Y CITAS",font=("Georgia",15,"bold"),fg="#1E2761").pack(pady=(15,5))
f4=tk.Frame(inner4); f4.pack(fill="x",padx=60,pady=5)

lbl(f4,"Codigo: *",1);               au_cod=ent(f4,1)
lbl(f4,"Tipo:",2);                   au_tip=tk.StringVar()
combo(f4,au_tip,["Audiencia Oral","Audiencia Preliminar","Cita con Cliente","Reunion Interna","Conciliacion"],2)
lbl(f4,"Caso Relacionado:",3);       au_cas_v=tk.StringVar()
au_cas_cb=ttk.Combobox(f4,textvariable=au_cas_v,width=26,state="readonly"); au_cas_cb.grid(row=3,column=1,sticky="w",pady=6)
lbl(f4,"Fecha y Hora:",4)
au_fec=DateEntry(f4,width=26,date_pattern="yyyy-mm-dd",font=("Calibri",11)); au_fec.grid(row=4,column=1,sticky="w",pady=6)
lbl(f4,"Duracion Estimada:",5);      au_dur=ent(f4,5)
lbl(f4,"Lugar: *",6);                au_lug=ent(f4,6)
lbl(f4,"Participantes Internos:",7); au_pin=ent(f4,7)
lbl(f4,"Participantes Externos:",8); au_pex=ent(f4,8)
lbl(f4,"Proposito:",9)
au_pro=tk.Text(f4,width=30,height=3,font=("Calibri",11),relief="solid",bd=1,bg="#FFFFFF")
au_pro.grid(row=9,column=1,sticky="w",pady=6)
lbl(f4,"Resultado Esperado:",10);    au_res=ent(f4,10)
lbl(f4,"Resultado Real:",11);        au_rea=ent(f4,11)

f4_tip=tk.StringVar()
f4_des4=[None]; f4_has4=[None]

def refresh_combos4():
    cas=get_casos_combo()
    au_cas_cb["values"]=list(cas.keys()); au_cas_cb._cas=cas

def get_d4():
    cas=getattr(au_cas_cb,"_cas",{})
    return {"codigo":au_cod.get().strip(),"tipo":au_tip.get(),
            "id_caso":cas.get(au_cas_v.get()),"fecha_hora":au_fec.get(),
            "duracion_estimada":au_dur.get().strip(),"lugar":au_lug.get().strip(),
            "participantes_int":au_pin.get().strip(),"participantes_ext":au_pex.get().strip(),
            "proposito":au_pro.get("1.0",tk.END).strip(),
            "resultado_esperado":au_res.get().strip(),"resultado_real":au_rea.get().strip()}

def refresh4():
    tree4.delete(*tree4.get_children())
    for r in sp_get_all_audiencias({"tipo":f4_tip.get() or None,
        "fecha_desde":f4_des4[0].get() if f4_des4[0] else None,
        "fecha_hasta":f4_has4[0].get() if f4_has4[0] else None}):
        tree4.insert("","end",values=(r["codigo"],r["tipo"] or "",r["numero_caso"] or "",
            r["fecha_hora"] or "",r["lugar"] or "",r["resultado_esperado"] or ""))

def on_sel4(e):
    s=tree4.selection()
    if not s: return
    r=sp_get_audiencia(str(tree4.item(s[0])["values"][0]))
    if not r: return
    for w,k in [(au_cod,"codigo"),(au_dur,"duracion_estimada"),(au_lug,"lugar"),
                (au_pin,"participantes_int"),(au_pex,"participantes_ext"),
                (au_res,"resultado_esperado"),(au_rea,"resultado_real")]:
        w.delete(0,tk.END); w.insert(0,r[k] or "")
    au_tip.set(r["tipo"] or "")
    au_pro.delete("1.0",tk.END); au_pro.insert("1.0",r["proposito"] or "")

def save4():
    d=get_d4(); e=validar_audiencia(d)
    if mostrar_errores(e): return
    try: sp_insert_audiencia(d); messagebox.showinfo("Guardado","Audiencia guardada correctamente."); clear4(); refresh4()
    except Exception as ex: messagebox.showerror("Error",str(ex))

def update4():
    if not messagebox.askyesno("Confirmar","Actualizar esta audiencia?"): return
    d=get_d4(); e=validar_audiencia(d)
    if mostrar_errores(e): return
    try: sp_update_audiencia(d); messagebox.showinfo("Actualizado","Audiencia actualizada correctamente."); refresh4()
    except Exception as ex: messagebox.showerror("Error",str(ex))

def delete4():
    cod=au_cod.get().strip()
    if not cod: messagebox.showwarning("Aviso","Selecciona una audiencia del listado."); return
    if not messagebox.askyesno("Eliminar",f"Eliminar '{cod}'?"): return
    try: sp_delete_audiencia(cod); messagebox.showinfo("Eliminado","Audiencia eliminada."); clear4(); refresh4()
    except Exception as ex: messagebox.showerror("Error",str(ex))

def clear4():
    for w in [au_cod,au_dur,au_lug,au_pin,au_pex,au_res,au_rea]: w.delete(0,tk.END)
    au_tip.set(""); au_cas_v.set(""); au_pro.delete("1.0",tk.END)

def excel4():
    rows=sp_get_all_audiencias({"tipo":f4_tip.get() or None,
        "fecha_desde":f4_des4[0].get() if f4_des4[0] else None,
        "fecha_hasta":f4_has4[0].get() if f4_has4[0] else None})
    exportar_excel("Audiencias y Citas",["Codigo","Tipo","N Caso","Fecha/Hora","Lugar","Res. Esperado","Res. Real"],
        [(r["codigo"],r["tipo"] or "",r["numero_caso"] or "",r["fecha_hora"] or "",
          r["lugar"] or "",r["resultado_esperado"] or "",r["resultado_real"] or "") for r in rows],
        f"Tipo:{f4_tip.get() or 'Todos'}")

def pdf4():
    rows=sp_get_all_audiencias({"tipo":f4_tip.get() or None,
        "fecha_desde":f4_des4[0].get() if f4_des4[0] else None,
        "fecha_hasta":f4_has4[0].get() if f4_has4[0] else None})
    exportar_pdf("Audiencias y Citas",["Codigo","Tipo","N Caso","Fecha/Hora","Lugar"],
        [(r["codigo"],r["tipo"] or "",r["numero_caso"] or "",r["fecha_hora"] or "",r["lugar"] or "") for r in rows],
        f"Tipo:{f4_tip.get() or 'Todos'}")

def build_filtros4(f):
    tk.Label(f,text="Tipo:",font=("Calibri",10)).grid(row=0,column=0,padx=5)
    ttk.Combobox(f,textvariable=f4_tip,values=["","Audiencia Oral","Audiencia Preliminar","Cita con Cliente","Reunion Interna","Conciliacion"],width=18,state="readonly").grid(row=0,column=1,padx=5)
    tk.Label(f,text="Desde:",font=("Calibri",10)).grid(row=0,column=2,padx=5)
    f4_des4[0]=DateEntry(f,width=12,date_pattern="yyyy-mm-dd"); f4_des4[0].grid(row=0,column=3,padx=5)
    tk.Label(f,text="Hasta:",font=("Calibri",10)).grid(row=0,column=4,padx=5)
    f4_has4[0]=DateEntry(f,width=12,date_pattern="yyyy-mm-dd"); f4_has4[0].grid(row=0,column=5,padx=5)
    tk.Button(f,text="Filtrar",font=("Calibri",9,"bold"),bg="#1E2761",fg="white",command=refresh4).grid(row=0,column=6,padx=10)

hacer_filtro(inner4,build_filtros4)
tree4=make_tree(inner4,("Codigo","Tipo","N Caso","Fecha/Hora","Lugar","Res. Esperado"))
tree4.bind("<<TreeviewSelect>>",on_sel4)
hacer_botones(inner4,{"Guardar":save4,"Actualizar":update4,"Eliminar":delete4,"Limpiar":clear4,"Excel":excel4,"PDF":pdf4})
refresh_combos4()
refresh4()

# ════════════════════════════════════════════════════════════════
root.mainloop()